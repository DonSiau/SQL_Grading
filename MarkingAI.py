import os
import pandas as pd
import requests
from flask import Flask, request, render_template, send_file, flash
import time
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers

app = Flask(__name__)
app.secret_key = "secret_key"  # Required to use Flask's flash messaging

# Use absolute path for UPLOAD_FOLDER
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

URL = "https://openrouter.ai/api/v1/chat/completions"
HEADERS = {
    "Authorization": f"Bearer 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", #School should use their own API key
    "Content-Type": "application/json"
}

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Request for the uploaded files
        studentans = request.files['studentans']
        suggestedans = request.files['suggestedans']
        
        if studentans.filename == '' or suggestedans.filename == '':
            return "No selected file", 400
        
        # Save the files requested into a folder
        studentans_filename = studentans.filename
        suggestedans_filename = suggestedans.filename
            
        studentans_path = os.path.join(app.config['UPLOAD_FOLDER'], studentans_filename)
        suggestedans_path = os.path.join(app.config['UPLOAD_FOLDER'], suggestedans_filename)   
        studentans.save(studentans_path)
        suggestedans.save(suggestedans_path)

        # Check if the uploaded files are valid Excel or CSV files
        if not (studentans_path.endswith(('.xlsx', '.csv')) and suggestedans_path.endswith(('.xlsx', '.csv'))):
            flash("One or both of the uploaded files are not valid Excel or CSV files.")
            return render_template('testform.html')

        try:
            # Process the uploaded Excel or CSV files
            if studentans_path.endswith('.xlsx'):
                stdans = pd.read_excel(studentans_path, engine='openpyxl') 
            else:
                stdans = pd.read_csv(studentans_path)
            
            if suggestedans_path.endswith('.xlsx'):
                suggestans = pd.read_excel(suggestedans_path, engine='openpyxl')
            else:
                suggestans = pd.read_csv(suggestedans_path)
        except ValueError as ve:
            flash("An error occurred while reading the files: " + str(ve))
            return render_template('testform.html')
        
        # Run the marking AI 
        processed_df = process_and_mark_answers(stdans, suggestans)

        # Create a file with the marked prefix
        output_filename = "marked_" + studentans_filename.rsplit('.', 1)[0] + ".xlsx"
        output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
            
        # Create a new workbook and add the data
        wb = openpyxl.Workbook()
        ws = wb.active

        for r in dataframe_to_rows(processed_df, index=False, header=True):
            ws.append(r)

        # Format the 'Total Score %' column as percentage
        total_score_col = processed_df.columns.get_loc('Total Score %') + 1  # +1 because openpyxl uses 1-based indexing
        for cell in ws[openpyxl.utils.get_column_letter(total_score_col)][1:]:
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

        # Save the formatted Excel file
        wb.save(output_filepath)

        # Return the marked version of the Excel sheet
        return send_file(output_filepath, as_attachment=True)
           
    return render_template('testform.html')

# Code to run the AI process
def process_and_mark_answers(stdans, suggestans):
    # Find columns that have the questions
    question_columns = [col for col in stdans.columns if col.startswith('Q')]

    # Add timestamp, student class, and name column into the marked data
    submissionData_columns = ['Timestamp', 'Enter your class', 'Enter your Student ID', 'Enter Your FULL Name']

    # Create columns for marks
    for col in question_columns:
        stdans[f'{col}_Mark'] = 0  

    max_score_per_question = 2  # Since the maximum score for each question is 2
    total_possible_marks = len(question_columns) * max_score_per_question  # Total possible score

    for index, row in stdans.iterrows():
        timestamp = row['Timestamp']
        student_class = row['Enter your class']
        student_ID = row['Enter your Student ID']
        student_Name = row['Enter Your FULL Name']
        for col in question_columns:
            # Extract student's answer
            answer = row[col]

            # Extract the question
            question_index = question_columns.index(col)

            # Extract the suggested answer from the suggested answers file
            suggested_answer = suggestans.iloc[0, question_index]
            
            # Prompt for the AI to check the answers
            prompt = ( 
                f"Suggested Answer: {suggested_answer},\n" 
                f"User Answer: {answer},\n\n" 
                "Evaluate the user's SQL answer based on the following criteria and be strict with it:\n"
                "1. If the user's answer is correct synthetically and has the same output as the suggested answer, return a score of 2.\n"
                "2. If the user's answer is an attempt but has errors or is partially correct, return a score of 1.\n"
                "3. If the user's answer does not have any text written in the string, return a score of 0.\n"
                "4. If there is a comment, follow it when marking.\n"
                "5. If there are multiple queries in the suggested answer, then the user's answers should include those queries (whilst following the other rules of course).\n"
                "Only return the score (2, 1, or 0) without any additional text."  
            )
            
            # Run the AI
            payload = {
                "model": "nousresearch/hermes-3-llama-3.1-405b",  # Make sure to use an instruct model, not a chat model https://openrouter.ai/
                "messages": [
                    {"role": "user", "content": prompt}
                ]
            }
            retries = 3
            delay = 2

            # Let the AI do multiple runs to ensure everything is marked and if there is an error there will be a -1
            for attempt in range(retries):
                try:
                    response = requests.post(URL, json=payload, headers=HEADERS)
                    result = response.json()
                    mark = int(result['choices'][0]['message']['content'].strip())
                    stdans.at[index, f'{col}_Mark'] = mark
                    break
                except Exception as e:
                    if attempt < retries - 1:  
                        time.sleep(delay)
                    else:
                        stdans.at[index, f'{col}_Mark'] = -1  

    spacer_col = 'marked->'
    stdans[spacer_col] = ''
    mark_columns = [col for col in stdans.columns if col.endswith('_Mark')]
    
    # Adding the marks to get the total mark
    stdans['total_marks'] = stdans[mark_columns].sum(axis=1)

    # Calculate the total percentage scored
    stdans['Total Score %'] = (stdans['total_marks'] / total_possible_marks)

    # Renaming the columns
    stdans = stdans.rename(columns={
        'Timestamp': 'Timestamp', 
        'Enter your class': 'Class', 
        'Enter your Student ID': 'StudentID',
        'Enter Your FULL Name': 'Name'
    })

    # Return the final data
    submissionData_columns = ['Timestamp', 'Class', 'StudentID', 'Name']
    stdans = stdans[submissionData_columns + question_columns + [spacer_col] + mark_columns + ['total_marks', 'Total Score %']]
    return stdans

if __name__ == '__main__':
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    app.run(host='0.0.0.0', port=5000, debug=True)